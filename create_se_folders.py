#!/usr/bin/env python3
"""
SE Folder Generator (Exact Match + Safety Checks)
Creates folders in SE Folder using EXACT strings from Excel column J
Example: A04_DR_8377191510_Huth Chandararith → folder name remains identical
"""
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ openpyxl not installed")
    print("   Run: uv pip install openpyxl")
    input("\nPress Enter to exit...")
    sys.exit(1)

def main():
    SCRIPT_DIR = Path(__file__).parent
    EXCEL_FILE = SCRIPT_DIR / "SE_List.xlsx"
    
    if not EXCEL_FILE.exists():
        print(f"❌ Missing Excel file: {EXCEL_FILE.name}")
        print("\nPlace your file here named 'SE_List.xlsx'")
        print("Or edit line 42 to match your actual filename")
        input("\nPress Enter to exit...")
        return
    
    print("="*60)
    print("📁 SE FOLDER GENERATOR (EXACT MATCH)")
    print("="*60)
    print(f"\nReading: {EXCEL_FILE.name}")
    print("Column J format: Area_Role_TelegramID_Name")
    print("Example: A04_DR_8377191510_Huth Chandararith\n")
    
    try:
        wb = load_workbook(EXCEL_FILE, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        print(f"❌ Excel error: {e}")
        input("\nPress Enter to exit...")
        return
    
    entries = []
    for row_idx, row in enumerate(ws.iter_rows(min_col=10, max_col=10, values_only=True), start=1):
        if row_idx == 1:
            continue
        
        cell_value = row[0]
        if not cell_value or not isinstance(cell_value, str):
            continue
        
        folder_name = cell_value.strip()
        # Skip if folder name would be invalid on Windows
        if folder_name.endswith(('.', ' ')):
            print(f"⚠️  Row {row_idx}: Skipped invalid name (trailing dot/space): '{folder_name}'")
            continue
        
        entries.append({"row": row_idx, "folder_name": folder_name})
    
    wb.close()
    
    if not entries:
        print("⚠️  No valid entries found in column J")
        print("   Check: Column J must contain values like 'A04_DR_8377191510_Huth Chandararith'")
        input("\nPress Enter to exit...")
        return
    
    print(f"✅ Found {len(entries)} entries:\n")
    for e in entries[:5]:
        print(f"   • {e['folder_name']}")
    if len(entries) > 5:
        print(f"   ... and {len(entries)-5} more")
    print()
    
    parent_dir = SCRIPT_DIR / "SE Folder"
    parent_dir.mkdir(exist_ok=True)
    print(f"📁 Creating folders inside: {parent_dir.name}\n")
    
    stats = {"created": 0, "existed": 0, "failed": 0, "conflict": 0}
    for e in entries:
        folder_path = parent_dir / e["folder_name"]
        
        if folder_path.exists():
            if folder_path.is_dir():
                print(f"   ⚠️  SKIPPED (exists): {e['folder_name']}")
                stats["existed"] += 1
            else:
                # Path exists but is a FILE (not folder) → conflict
                print(f"   ❌ CONFLICT: Path exists as FILE (not folder): {e['folder_name']}")
                stats["conflict"] += 1
                stats["failed"] += 1
        else:
            try:
                folder_path.mkdir()
                print(f"   ✅ CREATED: {e['folder_name']}")
                stats["created"] += 1
            except Exception as err:
                print(f"   ❌ FAILED: {e['folder_name']} → {err}")
                stats["failed"] += 1
    
    print("\n" + "="*60)
    print("📊 SUMMARY")
    print("="*60)
    print(f"Total entries processed : {len(entries)}")
    print(f"✓ Folders created       : {stats['created']}")
    print(f"⚠️  Already existed      : {stats['existed']}")
    if stats["conflict"] > 0:
        print(f"✗ Path conflicts        : {stats['conflict']} (file exists where folder should be)")
    print(f"✗ Failed                : {stats['failed']}")
    print(f"\n📁 Location: {parent_dir.resolve()}")
    input("\n✅ Done! Press Enter to close...")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⏹️  Stopped by user")
        input("Press Enter to exit...")
    except Exception as e:
        print(f"\n❌ CRITICAL ERROR: {e}")
        import traceback
        traceback.print_exc()
        input("\nPress Enter to exit...")